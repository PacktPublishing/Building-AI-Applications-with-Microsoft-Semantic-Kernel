import openpyxl
from typing_extensions import Annotated
from semantic_kernel.functions.kernel_function_decorator import kernel_function
import glob

class CheckSpreadsheet:

    @kernel_function(
        description="Checks that the spreadsheet contains the correct tabs, 2024 and 2025",
        name="CheckTabs",
    )
    def CheckTabs(self,
                  input: Annotated[str, "The path to the spreadsheet"]) -> Annotated[str, "The result of the check"]:
        if input.startswith("Error"):
            return input
        try:
            filePath = self.GetExcelFile(input)
            workbook = openpyxl.load_workbook(filePath)
            sheet_names = workbook.sheetnames
            if sheet_names == ['2024', '2025']:
                return input
            else:
                return "Error: the spreadsheet does not contain the correct tabs"
        except Exception as e:
            return f"Error: an exception {e} occurred when trying to open the spreadsheet"

    @kernel_function(
        description="Checks that the spreadsheet contains the correct cells A1-B5",
        name="CheckCells",
    )
    def CheckCells(self,
                 input: Annotated[str, "The path to the spreadsheet"]) -> Annotated[str, "The result of the check"]:
        if input.startswith("Error"):
            return input
        filePath = self.GetExcelFile(input)
        workbook = openpyxl.load_workbook(filePath)
        required_cells = {
            'A1': 'Quarter', 'B1': 'Budget',
            'A2': 'Q1', 'A3': 'Q2', 'A4': 'Q3', 'A5': 'Q4'
        }
        for year in ['2024', '2025']:          
            sheet = workbook[year]
            for cell, value in required_cells.items():
                if sheet[cell].value != value:
                    return "Error: missing quarters"
            for row in range(2, 6):
                if not isinstance(sheet[f'B{row}'].value, (int, float)):
                    return "Error: non-numeric inputs"
        return input

    @kernel_function(
        description="Checks that the spreadsheet contains the correct values, less than 1m per year and growth less than 10%",
        name="CheckValues",
    )
    def CheckValues(self,
                    input: Annotated[str, "The path to the spreadsheet"]) -> Annotated[str, "The result of the check"]:
        if input.startswith("Error"):
            return input
        filePath = self.GetExcelFile(input)
        workbook = openpyxl.load_workbook(filePath)
        years = ['2024', '2025']
        for year in years:
            if year not in workbook.sheetnames:
                return f"Error: Sheet for year {year} not found."
            sheet = workbook[year]
            values = [sheet[f'B{row}'].value for row in range(2, 6)]
            if not all(isinstance(value, (int, float)) for value in values):
                return f"Error: Non-numeric value found in sheet {year}."
            if sum(values) >= 1000000:
                return f"Error: Sum of values in year {year} exceeds 1,000,000."

            for i in range(len(values) - 1):
                if values[i + 1] > values[i] * 1.10:
                    return f"Error: More than 10% growth found from B{i+2} to B{i+3} in sheet {year}."
        return input
    
    def GetExcelFile(self, path: str) -> str:
        if path.startswith("Error"):
            return path
        try:
            excel_files = glob.glob(path + "/*.xlsx")
            if excel_files:
                return excel_files[0]
            else:
                return "Error: No Excel files found in the directory"
        except Exception as e:
            return f"Error: an exception {e} occurred when trying to open the spreadsheet"